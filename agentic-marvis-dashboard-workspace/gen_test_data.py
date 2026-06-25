import openpyxl

# --- 1. Project status workbook ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Status"
ws.append(["Metric", "Value"])
ws.append(["Overall Project Progress", "45%"])
ws.append(["AI Chatbot Completion", "65%"])
ws.append([])
ws.append(["Budget Item", "Amount"])
ws.append(["Total AI Dev Budget", 275000])
ws.append(["Spent To Date", 123750])
ws.append(["Remaining", 151250])
ws.append([])
ws.append(["Team", "Utilization %", "Notes"])
ws.append(["AI Team", 90, "Motivated but under pressure"])
ws.append(["IT Team", 68, "Improved server reliance"])
ws.append(["Support Team", 35, "Providing active chatbot feedback"])
ws.append([])
ws.append(["Risk", "Owner", "Status"])
ws.append(["Vendor SLA breach on model latency", "J. Tan", "At Risk"])
ws.append(["Headcount backfill delayed", "M. Reyes", "Watching"])
ws.append(["Security review pending sign-off", "K. Adams", "On Track"])
wb.save("test-data/project-status.xlsx")

# --- 2. Finance quarterly workbook ---
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "Finance"
ws2.append(["Metric", "Value"])
ws2.append(["Total Revenue", 13700000])
ws2.append(["Gross Profit", -1300000])
ws2.append(["Marketing Spend", 2000000])
ws2.append(["CAC", 48800])
ws2.append(["Active Customers", 88])
ws2.append([])
ws2.append(["Month", "Revenue", "COGS", "Gross Profit"])
months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
revs =  [610000,590000,640000,700000,730000,820000,910000,990000,1050000,1120000,1180000,1260000]
cogs =  [700000,680000,720000,760000,790000,860000,930000,1010000,1070000,1140000,1190000,1280000]
for m, r, c in zip(months, revs, cogs):
    ws2.append([m, r, c, r - c])
ws2.append([])
ws2.append(["Region", "Revenue"])
ws2.append(["North America", 5200000])
ws2.append(["APAC", 3100000])
ws2.append(["Europe", 2900000])
ws2.append(["LATAM", 2500000])
wb2.save("test-data/finance-quarterly.xlsx")

# --- 3. Sales / ops workbook ---
wb3 = openpyxl.Workbook()
ws3 = wb3.active
ws3.title = "Sales"
ws3.append(["Metric", "Value"])
ws3.append(["Total Sales", 15880])
ws3.append(["Customers", 20])
ws3.append(["Sales Goal", 16000])
ws3.append([])
ws3.append(["Category", "Revenue"])
ws3.append(["Handcrafted Soy Candle", 6200])
ws3.append(["Wax Melts", 4100])
ws3.append(["Diffusers", 3080])
ws3.append(["Gift Sets", 2500])
ws3.append([])
ws3.append(["Store", "Sales", "Status"])
ws3.append(["Downtown Flagship", 5200, "On Track"])
ws3.append(["Riverside Mall", 4100, "On Track"])
ws3.append(["Airport Kiosk", 1900, "At Risk"])
ws3.append(["Online Store", 4680, "On Track"])
wb3.save("test-data/sales-ops.xlsx")

print("done")
