#!/usr/bin/env python3
"""
Extract every table from an Excel workbook into structured JSON.

A "table" is a contiguous block of rows on a sheet, separated from other
blocks by blank rows. Most director-facing workbooks pack several small
tables onto one sheet (a KPI block, then a trend table, then a risk list) -
this splits them apart so each can be reasoned about independently instead
of as one ragged grid.

Usage:
    python extract_excel.py <path-to-xlsx> [--sheet SHEET_NAME] [--max-rows N] [--cache PATH]

`--cache PATH` additionally writes this run's JSON to PATH. Do this any
time you build a dashboard for a workbook the user might send an updated
version of later - the cached JSON is what diff_extract.py compares
against on the next run to catch renamed/added/removed columns instead of
re-deriving the dashboard from scratch and silently missing what changed.

Output: JSON on stdout. Shape:
{
  "source_file": "...",
  "sheets": [
    {
      "sheet_name": "...",
      "tables": [
        {
          "table_id": "Sheet1__t0",
          "anchor_cell": "A1",
          "headers": ["Month", "Revenue", "COGS"],
          "row_count": 12,
          "rows": [{"Month": "Jan", "Revenue": 120000, "COGS": 80000}, ...],
          "truncated": false,
          "columns": {
            "Month": {"inferred_type": "text", "non_null": 12},
            "Revenue": {"inferred_type": "currency", "non_null": 12,
                        "min": 80000, "max": 210000, "sum": 1640000, "mean": 136666.7},
            "COGS": {"inferred_type": "currency", ...}
          }
        }
      ]
    }
  ]
}

Type inference is heuristic (currency/percent/date/numeric/text) so the
caller can decide KPI vs. chart vs. table without re-deriving it by eye.
Numeric stats (min/max/sum/mean) are computed here, in Python, so any
number that ends up on the dashboard is traceable to a real calculation
rather than an eyeballed guess.
"""

import argparse
import json
import re
import sys
from datetime import datetime, date

import openpyxl


def infer_column_type(values):
    """Inspect non-null values in a column and guess a semantic type."""
    sample = [v for v in values if v is not None and v != ""]
    if not sample:
        return "empty"

    if all(isinstance(v, (datetime, date)) for v in sample):
        return "date"

    numeric_count = sum(isinstance(v, (int, float)) and not isinstance(v, bool) for v in sample)
    if numeric_count == len(sample):
        # percent heuristic: openpyxl gives 0-1 floats for cells formatted as %
        return "numeric"

    str_sample = [str(v) for v in sample]
    if all(re.match(r"^-?\$?[\d,]+\.?\d*%?$", s.strip()) for s in str_sample):
        if any("%" in s for s in str_sample):
            return "percent_text"
        if any("$" in s for s in str_sample):
            return "currency_text"
        return "numeric_text"

    return "text"


def column_stats(values, inferred_type):
    nums = [v for v in values if isinstance(v, (int, float)) and not isinstance(v, bool)]
    non_null = sum(1 for v in values if v is not None and v != "")
    stats = {"inferred_type": inferred_type, "non_null": non_null}
    if inferred_type == "numeric" and nums:
        stats.update(
            min=min(nums),
            max=max(nums),
            sum=round(sum(nums), 4),
            mean=round(sum(nums) / len(nums), 4),
        )
    return stats


def cell_value(cell):
    v = cell.value
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v


def split_into_tables(sheet):
    """Walk a sheet and split it into blocks separated by fully-blank rows."""
    rows = list(sheet.iter_rows())
    blocks = []
    current = []
    current_anchor = None

    def row_is_blank(row):
        return all(c.value is None for c in row)

    for row in rows:
        if row_is_blank(row):
            if current:
                blocks.append((current_anchor, current))
                current = []
                current_anchor = None
            continue
        if not current:
            current_anchor = row[0].coordinate
        current.append(row)
    if current:
        blocks.append((current_anchor, current))

    return blocks


def block_to_table(sheet_name, idx, anchor, rows, max_rows):
    # First row of the block is the header row.
    header_cells = rows[0]
    headers = []
    for i, c in enumerate(header_cells):
        h = c.value
        headers.append(str(h).strip() if h is not None else f"col_{i+1}")

    # Sheets often have a wider used-range than any single block actually
    # needs (a longer row elsewhere pads every block's iter_rows() width).
    # Drop trailing columns that are blank in the header AND every data row.
    while headers:
        last = len(headers) - 1
        header_blank = header_cells[last].value is None
        data_blank = all(
            (r[last].value is None if last < len(r) else True) for r in rows[1:]
        )
        if header_blank and data_blank:
            headers.pop()
            header_cells = header_cells[:-1]
        else:
            break

    data_rows_raw = rows[1:]
    truncated = len(data_rows_raw) > max_rows
    data_rows_raw = data_rows_raw[:max_rows]

    records = []
    columns_raw = {h: [] for h in headers}
    for r in data_rows_raw:
        record = {}
        for i, h in enumerate(headers):
            val = cell_value(r[i]) if i < len(r) else None
            record[h] = val
            columns_raw[h].append(val)
        if any(v is not None and v != "" for v in record.values()):
            records.append(record)

    columns = {}
    for h in headers:
        inferred = infer_column_type(columns_raw[h])
        columns[h] = column_stats(columns_raw[h], inferred)

    return {
        "table_id": f"{sheet_name}__t{idx}",
        "anchor_cell": anchor,
        "headers": headers,
        "row_count": len(records),
        "rows": records,
        "truncated": truncated,
        "columns": columns,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("path")
    parser.add_argument("--sheet", default=None, help="Only extract this sheet name")
    parser.add_argument("--max-rows", type=int, default=2000, help="Cap rows per table")
    parser.add_argument("--cache", default=None, help="Also write this run's JSON to PATH, for later diff_extract.py comparisons")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.path, data_only=True)
    sheet_names = [args.sheet] if args.sheet else wb.sheetnames

    out_sheets = []
    for name in sheet_names:
        if name not in wb.sheetnames:
            print(f"Sheet not found: {name}", file=sys.stderr)
            continue
        sheet = wb[name]
        blocks = split_into_tables(sheet)
        tables = [
            block_to_table(name, i, anchor, rows, args.max_rows)
            for i, (anchor, rows) in enumerate(blocks)
            if len(rows) >= 1
        ]
        out_sheets.append({"sheet_name": name, "tables": tables})

    result = {"source_file": args.path, "sheets": out_sheets}
    json.dump(result, sys.stdout, indent=2, default=str)

    if args.cache:
        with open(args.cache, "w", encoding="utf-8") as f:
            json.dump(result, f, indent=2, default=str)
        print(f"\nCached extraction to {args.cache}", file=sys.stderr)


if __name__ == "__main__":
    main()
