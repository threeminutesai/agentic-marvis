import openpyxl, json, re, sys
from datetime import datetime, date

def infer_column_type(values):
    sample = [v for v in values if v is not None and v != ""]
    if not sample:
        return "empty"
    if all(isinstance(v, (datetime, date)) for v in sample):
        return "date"
    numeric_count = sum(isinstance(v, (int, float)) and not isinstance(v, bool) for v in sample)
    if numeric_count == len(sample):
        return "numeric"
    str_sample = [str(v) for v in sample]
    if all(re.match(r"^-?\$?[\d,]+\.?\d*%?$", s.strip()) for s in str_sample):
        if any("%" in s for s in str_sample):
            return "percent_text"
        if any("$" in s for s in str_sample):
            return "currency_text"
        return "numeric_text"
    return "text"

def column_stats(values, inferred_type):
    nums = [v for v in values if isinstance(v, (int, float)) and not isinstance(v, bool)]
    non_null = sum(1 for v in values if v is not None and v != "")
    stats = {"inferred_type": inferred_type, "non_null": non_null}
    if inferred_type == "numeric" and nums:
        stats.update(min=min(nums), max=max(nums), sum=round(sum(nums), 4), mean=round(sum(nums)/len(nums), 4))
    return stats

def cell_value(cell):
    v = cell.value
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v

def row_is_blank(row):
    return all(c.value is None for c in row)

def split_into_tables(sheet):
    rows = list(sheet.iter_rows())
    blocks = []
    current = []
    current_anchor = None
    for row in rows:
        if row_is_blank(row):
            if current:
                blocks.append((current_anchor, current))
                current = []
                current_anchor = None
            continue
        if not current:
            current_anchor = row[0].coordinate
        current.append(row)
    if current:
        blocks.append((current_anchor, current))
    return blocks

def block_to_table(sheet_name, idx, anchor, rows, max_rows=2000):
    header_cells = rows[0]
    headers = []
    for i, c in enumerate(header_cells):
        h = c.value
        headers.append(str(h).strip() if h is not None else f"col_{i+1}")
    while headers:
        last = len(headers) - 1
        header_blank = header_cells[last].value is None
        data_blank = all((r[last].value is None if last < len(r) else True) for r in rows[1:])
        if header_blank and data_blank:
            headers.pop()
            header_cells = header_cells[:-1]
        else:
            break
    data_rows_raw = rows[1:]
    truncated = len(data_rows_raw) > max_rows
    data_rows_raw = data_rows_raw[:max_rows]
    records = []
    columns_raw = {h: [] for h in headers}
    for r in data_rows_raw:
        record = {}
        for i, h in enumerate(headers):
            val = cell_value(r[i]) if i < len(r) else None
            record[h] = val
            columns_raw[h].append(val)
        if any(v is not None and v != "" for v in record.values()):
            records.append(record)
    columns = {}
    for h in headers:
        inferred = infer_column_type(columns_raw[h])
        columns[h] = column_stats(columns_raw[h], inferred)
    return {
        "table_id": f"{sheet_name}__t{idx}",
        "anchor_cell": anchor,
        "headers": headers,
        "row_count": len(records),
        "rows": records,
        "truncated": truncated,
        "columns": columns,
    }

path = r"C:\Users\leona\Downloads\Dashboard_Financial_P&L_v3.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

out_sheets = []
for name in wb.sheetnames:
    sheet = wb[name]
    blocks = split_into_tables(sheet)
    tables = [
        block_to_table(name, i, anchor, rows)
        for i, (anchor, rows) in enumerate(blocks)
        if len(rows) >= 1
    ]
    out_sheets.append({"sheet_name": name, "tables": tables})

result = {"source_file": path, "sheets": out_sheets}
print(json.dumps(result, indent=2, default=str))
