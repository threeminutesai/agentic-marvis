import openpyxl, json, sys

path = r"C:\Users\leona\Downloads\Dashboard_Financial_P&L_v3.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

out = {}
for sname in wb.sheetnames:
    ws = wb[sname]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(v) if v is not None else None for v in row])
    out[sname] = rows

print(json.dumps(out, indent=2, ensure_ascii=False))
